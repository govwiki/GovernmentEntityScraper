import openpyxl
from pathlib import Path
import search
import sys
from bs4 import BeautifulSoup
import requests


def isValidUrl(url):
    success = False
    try:
        r = requests.get(url)
        success = (r.status_code == 200)
    except Exception as e:
        # print("error " + str(e))
        pass
    return success


xlsx_file = Path(sys.argv[1])
print(xlsx_file)
helper = search.Search()

# read file
wb_obj = openpyxl.load_workbook(xlsx_file)

# read active sheet from workbook
wsheet = wb_obj.active


rowNumber = 1
count = 0
for row in wsheet.iter_rows(max_row=wsheet.max_row):
    rowNumber += 1
    nameCell = 'A' + str(rowNumber)
    urlCell = 'Y' + str(rowNumber)
    entityName = wsheet[nameCell].value
    urlValue = wsheet[urlCell].value
    isValid = isValidUrl(urlValue)
    if not isValid:
        print(str(entityName) + " --- " + str(urlValue))
    if (rowNumber > 5000):
        break
