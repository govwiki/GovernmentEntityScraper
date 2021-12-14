from os import stat
import openpyxl
from pathlib import Path
from bs4 import BeautifulSoup
import requests
import logging
import datetime
import sys

# Setup logging
now = datetime.datetime.now()
date_time = now.strftime("%Y-%m-%d %H-%M-%S")

logger = logging.getLogger()
handler = logging.FileHandler(
    mode="w", filename="logs/url_checker_" + date_time + ".txt")
handler.setLevel(logging.INFO)
logger.addHandler(handler)
logger.setLevel(logging.INFO)

handler = logging.StreamHandler(sys.stdout)
handler.setLevel(logging.DEBUG)
logger.addHandler(handler)

"""
   Return the status code of the url using the 'requests' library.
   Returns -1 if there was an exception parsing
"""


def getStatusCode(url):
    success = -1
    try:
        r = requests.get(url, timeout=10)
        success = r.status_code
    except Exception as e:
        return -1
    return success


"""
 excel_filename -- an excel file where the active spreadsheet is a listing of Government entities
    with an "entity name" columns and a "url" column specified by the function parameters
     "entityNameColumn" and "entityUrlColumn" respectively.
"""


def getUrlResults(excel_filename, entityNameColumn, entityUrlColumn, header_exists=True, debug=False):
    xlsx_file = Path(excel_filename)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    wsheet = wb_obj.active
    rowNumber = 0
    if header_exists:
        rowNumber = 1
    statusCodeMap = {}
    collection = []
    for row in wsheet.iter_rows(max_row=wsheet.max_row):
        rowNumber += 1
        nameCell = entityNameColumn + str(rowNumber)
        urlCell = entityUrlColumn + str(rowNumber)
        entityName = wsheet[nameCell].value
        entityUrl = wsheet[urlCell].value

        statusCode = getStatusCode(entityUrl)
        statusCodeMap[statusCode] = statusCodeMap.get(statusCode, 0) + 1
        if debug:
            logging.info("%s -- %s (%s)" % (entityName, entityUrl, statusCode))

        entityUrlInfo = EntityUrlInfo(entityName, entityUrl, statusCode)
        collection.append(entityUrlInfo)

    if debug:
        logging.info("status code map: %s" % (statusCodeMap))
    return EntityUrlCollection(collection, statusCodeMap)


class EntityUrlCollection:
    def __init__(self, collection, map):
        self.collection = collection
        self.map = map


class EntityUrlInfo:
    def __init__(self, entityName, entityUrl, statusCode):
        self.entityName = entityName
        self.entityUrl = entityUrl
        self.statusCode = statusCode
