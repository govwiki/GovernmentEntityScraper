import sys
import openpyxl
import requests
from bs4 import BeautifulSoup
import time



def get_config(config_file):
    with open(config_file, 'r') as f:
        config = f.readlines()
        config = [c.rstrip() for c in config]
    return config


def find_on_page(query, headers):
    time.sleep(2)
    response = requests.get(query, headers=headers)
    urls_download = []
    print(response.status_code)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        for g in soup.find_all('div', 'tF2Cxc'):
            for site in g.find_all('div', 'yuRUbf'):
                for s in site.find_all('a'):
                    if not s.get('class') and s.get('data-jsarwt'):
                        urls_download.append(s['href'])
                        break

    return urls_download


def get_url(input_file, sheetname, columnnumber, outputfile, config_file, startRow=2, endRow=1000, year=2021):
    USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:65.0) Gecko/20100101 Firefox/65.0"
    #USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36"

    headers = {"user-agent": USER_AGENT}
    config = get_config(config_file)
    print(config)
    file = openpyxl.load_workbook(input_file)
    file_object = file[sheetname]

    links_pdf = {}
    #links_pdf = {r: [0, 0, 0],  }

    for r in range(int(startRow), int(endRow)):
        url = file_object.cell(row=r, column=int(columnnumber)).value
        if url:
            links_pdf[r] = []
            for schema in config:
                query = schema.replace('%URL%', url).replace('%YEAR%', str(year))
                URLS = [f"https://google.com/search?q={query}"]
                urls_download = []

                for u in URLS:
                    #print(f"{url=}")
                    urls_download += find_on_page(u, headers)
                    pass

                    print(urls_download)
                if urls_download:
                    links_pdf[r].append(urls_download[0])
                else:
                    links_pdf[r].append('null')

    print(f"{links_pdf=}")
    for l in links_pdf:
        for i, ob in enumerate(links_pdf[l]):
            file_object.cell(row=l, column=i+5).value = ob
    file.save(outputfile)



def main(argv):
    if len(argv) < 5:
        print("Not enough options. There must be at least 5 first fields or more.")
        print("For example: input_file, sheet_name, column_number, output_file, config_file, startRow=(default=2), endRow=(default=1000), year=(default=2021)")
        return

    input_file = argv[0]
    sheet_name = argv[1]
    column_number = argv[2]
    output_file = argv[3]
    config_file = argv[4]
    startRow = 2
    endRow = 6
    year = 2021

    for ar in argv:
        if 'startRow=' in ar:
            startRow = ar.replace('startRow=', '')
        elif 'endRow=' in ar:
            endRow = ar.replace('endRow=', '')
        elif 'year=' in ar:
            year = ar.replace('year=', '')

    print(input_file, sheet_name, input_file, sheet_name, column_number, output_file, config_file, startRow, endRow, year)

    get_url(input_file, sheet_name, column_number, output_file, config_file, startRow=startRow, endRow=endRow, year=year)



if __name__ == '__main__':
    #get_url("./data/Local Education Authority Web Addresses.xlsx", 'Sheet1', 4, "new.xlsx", "config.txt", startRow=2, endRow=6, year=2021)
    main(sys.argv[1:])


