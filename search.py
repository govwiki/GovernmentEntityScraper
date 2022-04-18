import urllib
import requests
from bs4 import BeautifulSoup
import re
import url_checker
import valid_urls
import overriden_entities
import logging
import datetime
import sys
import time
import getopt

import openpyxl
from openpyxl import Workbook
from pathlib import Path
import concurrent.futures

# Setup logging
now = datetime.datetime.now()
date_time = now.strftime("%Y-%m-%d %H-%M-%S")

logger = logging.getLogger()
handler = logging.FileHandler(
    mode="w", filename="logs/search_" + date_time + ".txt")
handler.setLevel(logging.INFO)
logger.addHandler(handler)
logger.setLevel(logging.INFO)

handler = logging.StreamHandler(sys.stdout)
handler.setLevel(logging.DEBUG)
logger.addHandler(handler)

abbrev = {}
abbrev['st'] = 'saint'
abbrev['ind'] = 'independent'
abbrev['sch'] = 'school'
abbrev['dist'] = 'district'
abbrev['mt'] = 'mount'

invalid_urls = ['wikipedia', 'facebook',
                'books.google', 'city-data', 'mapquest', 'manta', 'yellowpages']
valid_suffixes = ['home', 'index', 'main', 'en.html', 'default', 'about']
valid_entities = ['county', 'town', 'city']
isd_words = ['isd', 'independent', 'school']
valid_2_urls = valid_urls.get_valid_urls()
overriden_entity_map = overriden_entities.get_overriden_entities()

def getGoogleSearchResults(query):
    # desktop user-agent
    USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:65.0) Gecko/20100101 Firefox/65.0"

    query = query.lower()
    query_words = query.split(" ")
    temp = []
    for word in query_words:
        temp.append(expand_abbreviation(word))
    query_words = temp

    is_school_district = is_isd(query_words)

    query = query.replace(' ', '+')
    URL = f"https://google.com/search?q={query}"

    headers = {"user-agent": USER_AGENT}
    resp = requests.get(URL, headers=headers)

    if resp.status_code == 200:
        soup = BeautifulSoup(resp.content, "html.parser")
        results = []
        
        topResults = 6
        count = 0
        for g in soup.find_all('div', 'tF2Cxc'):
            anchors = g.find_all('a')
            missingDivs = g.find_all('div', 'TXwUJf')
            count += 1
            if count > topResults:
                break
            for anchor in anchors:
                if 'href' in anchor.attrs:
                    href = anchor['href']
                    title = get_h3_title(anchor.contents)
                    # if title != '':
                    #     logger.info("Title: " + str(title))
                    if '/search' not in href and 'http' in href:
                        if len(missingDivs) == 0 or is_school_district:
                            result = LinkResult(title, href)
                            results.append(result)
                        else:
                            logger.info("missing div: " + str(href))

        return results
    else:
        print("bad request: " + str(resp.headers))
        return []

class LinkResult:
    """
        A data type that contains two elements
        title : str
            The page title of the web page. This is what shows up as the link text in a Google search result.
        link : str
            The full url of the web page
    """
    def __init__(self, title, href):
        self.title = title
        self.link = href
    
    def __str__(self):
        return self.title + " : " + self.link

def get_h3_title(contents : list):
    """
        A helper method used to help parse the html of a google search result.
    """
    for content in contents:
        if content.name == 'h3':
            return str(content.contents[0])
    return ""

def getMatchingLink(query : str, access_url : bool) -> str:
    """ 
        Parameters
        --------------
        query : str
            a google search query for a given entity (e.g. "CITY OF GATESVILLE Texas")
        
        access_url : bool
            a flag indicating whether we should try to access the url, and check the status code
            and possibly the contents of the page or if we should just rely on the 
            url itself and the title displayed in the google search results
        
        Returns a string representing the link of the correct matching link
    """
    if query in overriden_entity_map:
        link = overriden_entity_map[query]
        logger.info("Overriden entity. Correct url is: " + str(link))
        return link

    results = getGoogleSearchResults(query)
    matchingLink = ""
    for result in results:
        logger.info("Candidate Result: " + str(result))
        url_checker.getStatusCode(result.link)
        if is_valid_link1(result, query):
            if not access_url:
                matchingLink = result.link
                break
            else:
                status = url_checker.getStatusCode(result.link)
                logger.info(str(result) + " --- " + str(status))
                if is_valid_status(status):
                    matchingLink = result.link
                    break
    if matchingLink == '':
        for result in results:
            if is_valid_link2(result.link):
                if not access_url:
                    matchingLink = result.link
                    break
                else:
                    status = url_checker.getStatusCode(result.link)
                    logger.info("2nd pass: " + result.link + " --- " + str(status))
                    if is_valid_status(status):
                        matchingLink = result.link
                        break
    logger.info("Chosen url: " + str(matchingLink))
    logger.info("   ")
    return matchingLink

def is_valid_status(status):
    return status == 200 or status == 406 or status == 403
    
def expand_abbreviation(word):
    w = word.lower()
    if w in abbrev:
        return abbrev[w]
    return w

def is_isd(queryWords : list) -> bool:
    """
        Does the Google search query pertain to an ISD (Independent School District)
    """
    ind = False
    sch = False
    isd_abbrev = False
    for word in queryWords:
        if word == 'independent':
            ind = True
        elif word == 'school':
            sch = True
        elif word == 'isd':
            return True
    return ind and sch

def contains_isd_words(title : str):
    """
        Does the page title of a web page contain keywords that indicate it is likely an ISD
    """
    for title_word in title.split(" "):
        for isd_word in isd_words:
            if isd_word in title_word.lower():
                return True
    return False

def is_valid_link1(result : LinkResult, query : str) -> bool:
    """
        Parameters
        --------------
        result : LinkResult
           a link result object that contains the page title and the url
        
        query : str
           the google search query used to retrieve this result
        
        Returns true if the result is a valid result for the given search query.
        First pass of algorithm.
    """
    link = result.link
    title = result.title

    slashes = num_slashes(link)
    words = query.split(" ")
    new_words = []
    for word in words:
        new_words.append(expand_abbreviation(word.lower()))
    
    words = new_words
    
    if len(words) >= 3:
        entityType = words[0].lower()
        if entityType in valid_entities:
            # Check if city, town, county is in text
            entityName = expand_abbreviation(words[2].lower())
            if entityName not in title.lower():
                return False
            if len(words) >= 5:
                entityName2 = words[3].lower()
                if entityName2 not in title.lower():
                    return False
        else:
            # Special district
            nums = "123456789"
            if words[-2] in nums:
                number = words[-2].lower()
                if number not in title.lower():
                    return False
            x, y = words[0].lower(), words[1].lower()
            if x not in title.lower() and x not in link.lower():
                return False
    
    # Check if a valid school district by consulting the title of the page
    if is_isd(words):
        if not contains_isd_words(title):
            return False

    if slashes >= 4:
        return False
    if slashes == 3:
        # The page is one level deep; check the list of hardcoded valid pages
        suffix = getPath(link)
        for valid_suffix in valid_suffixes:
            if valid_suffix in suffix:
                return True
        return False
    for invalid_url in invalid_urls:
        if invalid_url in link:
            return False
    return True


def is_valid_link2(url):
    """
        2nd pass of the algorithm where we check if the url is valid.
        basically, we allow all web pages (with arbitrary path depth) 
        that originate from "valid_2_urls" list.
    """
    slashes = num_slashes(url)
    for valid_url in valid_2_urls:
        if valid_url in url:
            return True
    return False

def num_slashes(url):
    """ Count the number of slashes in a url, ignoring the last character """
    count = 0
    for letter in url[:-1]:
        if letter == '/':
            count += 1
    return count


def getPath(url):
    """
        Returns the path portion of a url. I.e. everything that comes after the hostname.
        Done this by counting the number of slashes
    """
    count = 0
    for i, letter in enumerate(url[:-1]):
        if letter == '/':
            count += 1
        elif count >= 3:
            return url[i:]
    return ''


def iterate(excel_filename, tab_name, column, output_file=None, fn=None,
            suffix="",
            header_exists=True, debug=False,
            startRow=1, parallel=True, match_correct=False, endRow=1,
            access_url=True):
    xlsx_file = Path(excel_filename)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    wsheet = wb_obj[tab_name]
    rowNumber = 1
    if header_exists:
        rowNumber = 2
    statusCodeMap = {}
    collection = []
    if(Path(output_file).exists()):
        book = openpyxl.load_workbook(output_file)
    else:
        book = Workbook()
        book.create_sheet()
    increment = 10
    start = time.time()
    if parallel:
        """ Parallel execution of requests. 
            Although faster than non-parallel variant, we've experimentally found that this
            triggers Google to block subsequent results due to high load.
        """
        while rowNumber < startRow:
            rowNumber += 1

        while rowNumber < wsheet.max_row and rowNumber < endRow:
            entities = getEntities(
                rowNumber, wsheet, increment, column, suffix)
            with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                for index, (entityName, entityUrl) in enumerate(zip(entities, executor.map(fn, entities))):
                    book['Sheet1']['A' +
                                   str(index+rowNumber)].value = entityName
                    book['Sheet1']['B' +
                                   str(index+rowNumber)].value = entityUrl
                    # print("%s -- %s" % (entityName, entityUrl))
            book.save(output_file)
            elapsed = time.time() - start
            print("%d  (%.1f)" % (rowNumber, elapsed))
            rowNumber += increment
    else:
        if match_correct:
            """ This flag is used to generate urls for entities that have a ground truth url 
               (i.e. column C 'correct url' is populated with a url) 
            """
            count = 0
            num_correct = 0
            num_filled = 0
            saved = False
            for row in wsheet.iter_rows(max_row=wsheet.max_row):
                rowNumber += 1
                if rowNumber < startRow:
                    continue

                nameCell = column + str(rowNumber)
                if wsheet[nameCell].value is None:
                    continue
                entityName = wsheet[nameCell].value + " " + suffix

                correct_url = book['Sheet1']['C' + str(rowNumber)].value
                if correct_url is not None:
                    logger.info("   ")
                    logger.info(entityName + ": \"" + str(correct_url) + "\"")
                    res = fn(entityName, access_url)
                    book['Sheet1']['A' + str(rowNumber)].value = entityName
                    book['Sheet1']['B' + str(rowNumber)].value = res
                    if res == correct_url or (res == '' and correct_url == 'None'):
                        book['Sheet1']['D' + str(rowNumber)].value = 'Yes'
                        num_correct += 1
                        num_filled += 1
                    elif (res == '' and correct_url != 'None'):
                        book['Sheet1']['D' + str(rowNumber)].value = 'Blank'
                    else:
                        book['Sheet1']['D' + str(rowNumber)].value = 'No'
                        num_filled += 1
                    count += 1
                    saved = False
                if count % 10 == 0 and not saved:
                    elapsed = time.time() - start
                    logger.info("%d correct: %d, filled: %d, total: %d",
                                rowNumber, num_correct, num_filled, count)
                    logger.info("accuracy: %.3f " %
                                (num_correct / (num_filled + 0.01)))
                    book.save(output_file)
                    saved = True

            logger.info("%d correct: %d, filled: %d, total: %d",
                        rowNumber, num_correct, num_filled, count)
            book.save(output_file)

        else:
            """ This code snippet is used to generate urls for all entities """
            for row in wsheet.iter_rows(max_row=wsheet.max_row):
                rowNumber += 1
                if rowNumber < startRow:
                    continue
                nameCell = column + str(rowNumber)
                entityName = wsheet[nameCell].value + " " + suffix
                logger.info("   ")
                logger.info(entityName)

                res = fn(entityName, access_url)
                book['Sheet1']['A' + str(rowNumber)].value = entityName
                book['Sheet1']['B' + str(rowNumber)].value = res
                if rowNumber % 10 == 0:
                    elapsed = time.time() - start
                    logger.info("%d (%.1f)" % (rowNumber, elapsed))
                    book.save(output_file)

        book.save(output_file)


def getEntities(start_row, wsheet, increment, column, suffix):
    max_row = min(start_row+increment, wsheet.max_row)
    entities = []
    for row in range(start_row, max_row+1, 1):
        nameCell = column + str(row)
        entityName = wsheet[nameCell].value + " " + suffix
        entities.append(entityName)
    return entities


### Main method
# iterate('data/Govt_Units_2017_Final.xlsx', 'General Purpose',
#         'B', output_file='data/govt_units_18_04_22.xlsx', fn=getMatchingLink,
#         suffix="",
#         debug=True, parallel=False, match_correct=False, startRow=1, endRow=1000, access_url=False)

def main(argv):
    inputfile = None
    outputfile = None
    sheetname = None
    columnname = None
    try:
        opts, args = getopt.getopt(argv, "hi:o:s:c:", ["ifile=", "ofile=", "sheetname=", "columnname="])
    except getopt.GetoptError:
        print
        'search.py -i <inputfile> -o <outputfile> -s <sheetname> -c <columnname>'
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print
            'test.py -i <inputfile> -o <outputfile> -s <sheetname> -c <columnname>'
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg
        elif opt in ("-s", "--sheetname"):
            sheetname = arg
        elif opt in ("-c", "--columnname"):
            columnname = arg
    print
    'Input file is "', inputfile
    print
    'Sheet name is "', sheetname
    print
    'Column name is "', columnname
    print
    'Output file is "', outputfile
    iterate(inputfile, sheetname,
            columnname, outputfile, fn=getMatchingLink,
            suffix="",
            debug=True, parallel=False, match_correct=False, startRow=1, endRow=1000, access_url=False)


if __name__ == "__main__":
   main(sys.argv[1:])

# iterate('data/Texas Local Governments.xlsx', 'Census of Govts',
#         'D', output_file='data/texas_websites_01_10_22.xlsx', fn=getMatchingLink,
#         suffix="Texas",
#         debug=True, parallel=False, match_correct=False, startRow=1, endRow=6000, access_url=False)
